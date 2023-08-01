import os
from string import ascii_uppercase
from typing import Iterable, Any

import openpyxl
from openpyxl.cell import Cell
from openpyxl.styles import Alignment, Font, Protection
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.worksheet.datavalidation import DataValidation
from loguru import logger

from . import config
from .exceptions import RuntimeException


def prepare_excel_file(filename: str) -> openpyxl.Workbook:
	"""
	Подготовка основного excel-файла.
	"""
	file = read_excel_file(filename)
	cleaned_up_wb = cleanup_file_cells(wb=file)
	if not os.path.exists(config.OUTPUT_PATH):
		os.mkdir(config.OUTPUT_PATH)
	cleaned_up_wb.save(filename=f"{config.OUTPUT_PATH}/{filename}")

	logger.info("Успешно произведена очистка таблицы от лишних символов в ячейках. "
		  f"Файл записан в папку '{config.OUTPUT_PATH}'.")

	return cleaned_up_wb


def read_excel_file(filepath: str) -> openpyxl.Workbook:
	"""
	Чтение основного файла и проверка его содержания.
	"""
	try:
		file: openpyxl.Workbook = openpyxl.load_workbook(filename=filepath, data_only=True)
		if not all(
			(sheet in (str(sh).lower() for sh in file.sheetnames) for sheet in config.NEEDED_SHEETS.values())
		):
			raise RuntimeException("Not all searched sheets was founded.\n"
							   f"Searching for sheets: {list(config.NEEDED_SHEETS.values())}")
	except Exception as e:
		raise RuntimeException(f"Can't read the Excel file. Exception: {e}")

	check_worksheets(file)

	logger.info(f"Excel-файл '{filepath}' успешно обнаружен и инициализирован.")

	return file


def check_worksheets(wb: openpyxl.Workbook) -> None:
	"""
	Проверяет листы на наличие столбцов, указанных в конфиге.
	"""
	main_worksheet: Worksheet = wb[config.NEEDED_SHEETS["MAIN_SHEET"]]
	secondary_worksheet: Worksheet = wb[config.NEEDED_SHEETS["SECONDARY_SHEET"]]

	def get_sheet_titles(sheet: Worksheet) -> Iterable[str]:
		sheet_titles_row = next(sheet.iter_rows())
		return (cell.value for cell in sheet_titles_row)

	def check_titles(columns: Iterable[str], searched_columns: Iterable[str], sheet_name: str) -> None:
		if any(
			(col not in searched_columns for col in tuple((str(item).strip() for item in columns)))
		):
			raise RuntimeException(f"Got an unexpected columns names in sheet '{sheet_name}'.\n"
							   f"Expected columns: {list(searched_columns)}")

	main_sheet_titles = get_sheet_titles(main_worksheet)
	secondary_sheet_titles = get_sheet_titles(secondary_worksheet)

	check_titles(
		columns=main_sheet_titles, searched_columns=config.MAIN_SHEET_COLUMNS.values(),
		sheet_name=config.NEEDED_SHEETS["MAIN_SHEET"]
	)
	check_titles(
		columns=secondary_sheet_titles, searched_columns=config.SECONDARY_SHEET_COLUMNS.values(),
		sheet_name=config.NEEDED_SHEETS["SECONDARY_SHEET"]
	)


def cleanup_file_cells(wb: openpyxl.Workbook) -> openpyxl.Workbook:
	"""
	На всякий случай удаляю лишние пробелы в ячейках.
	Можно добавить еще какое-нибудь другое форматирование.
	"""
	worksheets = wb.sheetnames
	for sheet in worksheets:
		ws = wb[sheet]
		for row in ws:
			for cell in row:
				if isinstance(cell.value, str):
					cell.value = cell.value.strip()
	return wb


def create_excel_file(sheet_name: str) -> openpyxl.Workbook:
	"""
	Создание Excel-файла
	"""
	wb = openpyxl.Workbook()
	sheet = wb.active
	sheet.title = sheet_name
	sheet.append(config.OUTPUT_FIELDS)
	return wb


def format_sheets(wb: openpyxl.Workbook, add_protection_and_validation: bool) -> openpyxl.Workbook:
	"""
	Форматирование Excel-файла
	"""
	alignment = Alignment(horizontal="center", vertical="center")
	font_first_row = Font(bold=True, size=12)

	for sheet_name in wb.sheetnames:
		sheet: Worksheet = wb[sheet_name]
		first_row = next(sheet.iter_rows())

		if add_protection_and_validation:
			sheet = add_sheet_protection(sheet, first_row)  # добавить защиту от изменений и валидацию данных

		for cell in first_row:
			cell.font = font_first_row  # добавить шрифт

		dims: dict = {}
		for row in sheet.rows:
			for cell in row:
				cell: Cell
				if cell.value:
					dims[cell.column_letter] = max((dims.get(cell.column_letter, 0), len(str(cell.value))))

				cell.alignment = alignment  # выравнивание текста

		for col, value in dims.items():
			sheet.column_dimensions[col].width = value + 5  # изменить ширину столбцов

	return wb


def add_sheet_protection(ws: Worksheet, first_row: list[Cell]) -> Worksheet:
	"""
	Устанавливает защиту на изменение Excel-файла.
	Но делает возможными для редактирования столбцы, которые указаны как таковые в конфиге.
	"""
	not_protected_columns = []
	columns_aliases_for_data_validation = {}
	for cell in first_row:
		try:
			cell_alias = next(key for key, val in config.MAIN_SHEET_COLUMNS.items() if val == cell.value)
			if cell_alias in config.PROTECTION_SHEETS_PARAMS["NOT_PROTECTED_COLUMNS"]:
				not_protected_columns.append(cell.column_letter)  # определяю столбцы, которые можно редактировать
			if cell_alias in config.CELLS_DATA_VALIDATION:
				validation_type = config.CELLS_DATA_VALIDATION.get(cell_alias)
				columns_aliases_for_data_validation.setdefault(cell_alias, validation_type)
		except StopIteration:
			continue  # В Output fields есть другие столбцы. Потом мб нужно будет поправить костыль.

	ws.protection.sheet = True

	for row in ws:
		for cell in row:
			cell: Cell
			if cell.column_letter in not_protected_columns:
				cell.protection = Protection(locked=False)

	if any(columns_aliases_for_data_validation):
		columns_letters_and_data_validation_types = {}
		for col_alias in columns_aliases_for_data_validation:
			validation_type = columns_aliases_for_data_validation.get(col_alias)
			col_letter = next(c.column_letter for c in first_row if c.value == config.MAIN_SHEET_COLUMNS.get(col_alias))
			columns_letters_and_data_validation_types.setdefault(col_letter, validation_type)
		ws = add_cells_validation(ws=ws, columns_letters_and_validation_types=columns_letters_and_data_validation_types)

	return ws


def add_cells_validation(ws: Worksheet, columns_letters_and_validation_types: dict[str, str]) -> Worksheet:
	"""
	Добавляет валидацию данных для столбцов.
	"""
	dv = None

	for col in columns_letters_and_validation_types:
		validation_type = columns_letters_and_validation_types.get(col)
		if validation_type == "integer":
			dv = DataValidation(type="decimal", operator="greaterThanOrEqual",
								formula1=0, showErrorMessage=True, showDropDown=True)
			dv.error = "Введите целое число больше или равное нулю"
			dv.errorTitle = "Неверная сумма"
			dv.prompt = dv.error
			dv.promptTitle = "Ввод суммы"
		if dv:
			ws.add_data_validation(dv)
			for row in ws:
				for cell in row:
					cell: Cell
					if cell.column_letter == col:
						dv.add(cell)

	return ws


def excel_file_to_list_of_dicts(ws: Worksheet, cut: bool = True) -> list[dict[str, Any]]:
	"""
	Форматирует Excel-файл в список словарей (привычный формат данных).
	"""
	first_row_number = 1
	rows = []
	for row in ws:
		row_data = {}
		for cell in row:
			cell: Cell
			cell_column_letter = ascii_uppercase[cell.column - 1]
			cell_column_name = ws[f"{cell_column_letter}{first_row_number}"].value
			row_data.setdefault(cell_column_name, cell.value)
		rows.append(row_data)
	if cut:
		return rows[1:-1]  # первая строка - заглавная, последняя - пустая ("итоговая")
	return rows[1:]  # если нет итоговой строки
