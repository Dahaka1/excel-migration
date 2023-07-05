from datetime import date
from string import ascii_uppercase
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.cell import Cell
from openpyxl.worksheet.worksheet import Worksheet
from loguru import logger

from . import config
from .models import Worker, WorkingPosition, Manager


def excel_file_to_list_of_dicts(ws: Worksheet) -> list[dict[str, Any]]:
	"""
	Форматирует Excel-файл в список словарей (привычный формат данных).
	"""
	first_row_number = 1
	rows = []
	for row in ws:
		row_data = {}
		for cell in row:
			cell: Cell
			cell_column_letter = ascii_uppercase[cell.column - 1]
			cell_column_name = ws[f"{cell_column_letter}{first_row_number}"].value
			row_data.setdefault(cell_column_name, cell.value)
		rows.append(row_data)
	return rows


def read_workers(wb: Workbook) -> list[Worker]:
	"""
	Считывает строки из файла.
	Возвращает список работников без определенной даты выплаты за текущий месяц.
	"""
	ws_main_name = next(ws for ws in wb.sheetnames if str(ws).lower() == config.NEEDED_SHEETS["MAIN_SHEET"])
	ws: Worksheet = wb[ws_main_name]
	rows = excel_file_to_list_of_dicts(ws=ws)
	rows = rows[1:-1]  # первая строка - заглавная, последняя - пустая ("итоговая")

	workers_list = [Worker(**row) for row in rows]
	workers_output = tuple(worker for worker in workers_list if worker.name != config.NULL_VALUE)

	logger.info(f"Работников с определенным ФИО обнаружено в исходном файле: {len(workers_output)}")

	ws_secondary_name = next(ws for ws in wb.sheetnames if str(ws).lower() == config.NEEDED_SHEETS["SECONDARY_SHEET"])

	return get_workers_paying_dates(
		workers=workers_output, ws=wb[ws_secondary_name]
	)


def get_workers_paying_dates(workers: tuple[Worker], ws: Worksheet) -> list[Worker]:
	"""
	Определяет дату выплаты для работника в текущем месяце (по побочному листу).
	Возвращает генератор с работниками с определенными датами.
	"""
	sheet_data = excel_file_to_list_of_dicts(ws=ws)
	current_month = config.MONTHS[date.today().month - 1]
	counter = 0
	for item in sheet_data:
		if not counter == 0:  # первую строку не надо считывать - заглавная
			del item[config.SECONDARY_SHEET_COLUMNS.get("PAYING_PERIOD")]  # не нужен этот столбец
			working_position_name = item.pop(config.SECONDARY_SHEET_COLUMNS.get("DISTRICT_COLUMN"))
			working_position = WorkingPosition(
				name=working_position_name, **item
			)
			for worker in workers:
				if worker.working_position.lower().strip() == working_position.name.lower().strip():
					worker.paying_date = working_position[current_month] or config.NULL_VALUE
		counter += 1

	return list(workers)


def write_files(workers: Iterable[Worker]) -> None:
	"""
	Создает все нужные итоговые файлы.

	Можно объединить классы в один общий или создать родительский, чтобы упростить код.
	Но сейчас пока пофиг :)
	"""
	all_managers_names = set(((worker.manager_name, worker.district) for worker in workers))

	workers_handled_while_writing_managers_files = 0
	handled_managers = []

	for manager_name, district in all_managers_names:
		manager_workers = [worker for worker in workers if worker.manager_name == manager_name]
		manager = Manager(name=manager_name, business_unit=district)
		manager.add_workers(
			*manager_workers
		)

		output_dir = f"{config.OUTPUT_PATH}/{district}"

		if district != config.NULL_VALUE:
			written_strings_amount = manager.write_file(output_dir=output_dir)
			workers_handled_while_writing_managers_files += written_strings_amount
		handled_managers.append(manager)

	workers_handled_while_writing_districts_files = Manager.write_districts_files(
		managers=handled_managers, workers=workers
	)

	report_text = f"Файлы по каждому менеджеру и каждому региону успешно отсортированы и записаны. " \
				  f"Всего в файлах по менеджерам работников определено: {workers_handled_while_writing_managers_files}\n"

	if workers_handled_while_writing_managers_files > workers_handled_while_writing_districts_files:
		report_text += "Конечное количество обработанных работников больше, ибо у некоторых " \
					   "менеджеров определены несколько регионов - файлы дублируются в папку каждого региона."
	logger.info(report_text)
