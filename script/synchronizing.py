import os
import time

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell import Cell
from loguru import logger

from . import config
from .models import Worker
from .exceptions import RuntimeException, ValueException
from .data import excel_file_to_list_of_dicts, format_sheets
from .utils import initiate_action


def update_workers_paying_amount(workers: tuple[Worker], updating_wb: Workbook, output_filename: str) -> None:
	"""
	Обновление суммы для работников в основном файле.
	"""
	managers_names = set(worker.manager_name for worker in workers)
	districts_names = set(worker.district for worker in workers)
	if config.NULL_VALUE in districts_names:
		logger.warning(f"В основном файле найдены работники со значением региона '{config.NULL_VALUE}' "
					   f"(неопределенный регион). Их данные не будут обновлены!")
	if config.SEVERAL_DISTRICTS_MANAGER_FILES_PATH in os.listdir(config.OUTPUT_PATH):
		districts_names.add(config.SEVERAL_DISTRICTS_MANAGER_FILES_PATH)
	updated_workers_result = []
	for district_name in districts_names:
		if district_name != config.NULL_VALUE:
			district_path = f"{config.OUTPUT_PATH}/{district_name}"
			if not os.path.exists(district_path):
				raise RuntimeException(f"District '{district_name}' folder not found")
			for excel_filename in os.listdir(district_path):
				filename = excel_filename.replace(".xlsx", str())
				if filename != district_name:
					if filename not in managers_names:
						raise ValueException(f"Manager '{filename}' (filename '{excel_filename}') "
											 f"not found in managers list")
					manager_file_path = f"{district_path}/{excel_filename}"
					manager_workers = filter(lambda worker: worker.manager_name == filename, workers)
					updated_workers = update_data_from_manager_file(manager_file_path, tuple(manager_workers), updating_wb,
												  filename)
					updated_workers_result.extend(updated_workers)

	if any(updated_workers_result):
		fileout_path = f"{config.OUTPUT_PATH}/{output_filename}"
		updating_wb = format_sheets(updating_wb, add_protection_and_validation=False)
		updating_wb.save(fileout_path)
		logger.info(f"Сотрудников было успешно обновлено в основном файле (суммы выплат): "
					f"{len(updated_workers_result)}. Итоговый файл находится в директории '{config.OUTPUT_PATH}'.")
		time.sleep(1)  # иначе строки чет путаются порядком
		advanced_info = initiate_action(
			f"Нажмите '{config.ADVANCED_SYNC_VIEW_BUTTONS[0]}', чтобы получить расширенные данные об обновлении."
			f"\nИли нажмите \"Enter\", чтобы выйти.",
			config.ADVANCED_SYNC_VIEW_BUTTONS
		)
		if advanced_info:
			print("\n" + Worker.get_updating_info(updated_workers_result))
		else:
			exit(0)
	else:
		logger.info("Обновления данных по сотрудникам не найдены. Итоговый файл не был изменен.")


def update_data_from_manager_file(file_path: str, workers: tuple[Worker], main_file: Workbook,
								  manager_name: str) -> list[Worker]:
	"""
	Поиск данных для обновления и строки с данными менеджера в основном файле.
	Можно красиво было сделать через ООП (класс работника), но пока задача только одна и очень конкретная,
	 быстрей сделаю в лоб =)
	"""
	manager_file: Workbook = load_workbook(file_path, data_only=True, read_only=True)
	manager_last_name = manager_name.split()[0]
	manager_file_worksheet = manager_file[config.MANAGER_AND_DISTRICTS_MAIN_SHEET_NAME % manager_last_name]
	manager_file_data = excel_file_to_list_of_dicts(manager_file_worksheet, cut=False)
	main_file_worksheet = main_file[config.NEEDED_SHEETS["MAIN_SHEET"]]
	main_file_worksheet_data = excel_file_to_list_of_dicts(main_file_worksheet)
	updated_workers = []
	for row in manager_file_data:
		worker = Worker(**row)
		if worker.name not in (w.name for w in workers):
			raise ValueException(f"Worker '{worker.name}' (manager file '{file_path}')"
								 f" not found in general file")
		if worker.paying_amount and worker.paying_amount != config.NULL_VALUE:
			worker_main_file_data = next(r for r in main_file_worksheet_data if
									 r[config.MAIN_SHEET_COLUMNS["EMPLOYEE_COLUMN"]] == worker.name)
			current_worker_paying_amount = worker_main_file_data[config.MAIN_SHEET_COLUMNS["PAYING_AMOUNT"]]
			if not current_worker_paying_amount and worker.paying_amount or \
				int(current_worker_paying_amount) != int(worker.paying_amount):
				worker_row_number = worker_main_file_data.get(config.MAIN_SHEET_COLUMNS["INDEX_COLUMN"])
				write_worker_data(worker, int(worker_row_number), main_file_worksheet)
				updated_workers.append(worker)
	return updated_workers


def write_worker_data(worker: Worker, row_number: int, ws: Worksheet) -> None:
	"""
	Запись обновленных данных в файл (если были изменены).
	"""
	first_row = next(ws.iter_rows())
	for cell in first_row:
		cell: Cell
		if cell.value == config.MAIN_SHEET_COLUMNS["INDEX_COLUMN"]:
			index_col_letter = cell.column_letter
		elif cell.value == config.MAIN_SHEET_COLUMNS["PAYING_AMOUNT"]:
			paying_amount_col_letter = cell.column_letter
	for row in ws:
		if row not in [next(ws.iter_rows()), tuple(ws.iter_rows())[-1]]:  # первая строка - титульная, последняя - итоговая
			for cell in row:
				if cell.column_letter == index_col_letter and int(cell.value) == row_number:
					for cell_ in row:
						if cell_.column_letter == paying_amount_col_letter:
							cell_.value = int(worker.paying_amount)

