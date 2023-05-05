from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet

import data

main_file_name, file_path = data.config()


def read_file(sheet_name: str) -> Worksheet:
	workbook = load_workbook(f'{file_path}/{main_file_name}')
	workbook.active = workbook[sheet_name]
	worksheet = workbook.active
	return worksheet


def read_subfile(filepath):
	workbook = load_workbook(filepath)
	worksheet = workbook.active
	return workbook, worksheet
